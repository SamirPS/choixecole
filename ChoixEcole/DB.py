#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 12 14:19:23 2019
@author: robinchaussemy
"""
import openpyxl
import sqlite3

connexion = sqlite3.connect("choixecole.db")
curseur = connexion.cursor()


""" Retourne la taille du fichier et les valeurs"""

book = openpyxl.load_workbook('BdDEcoles.xlsx')
sheet = book.get_sheet_by_name('Ecoles')

"""Connaitre le nombre de colonnes et de lignes des feuilles """
r = sheet.max_row
c = sheet.max_column

def ecole():

    curseur.execute("DROP TABLE EcoleS")
    connexion.commit()
    curseur.execute("CREATE TABLE `EcoleS` (`Id`	INTEGER PRIMARY KEY AUTOINCREMENT,`Acronyme`	TEXT,`Nom`	TEXT,`Groupe`	TEXT,`Commune`	TEXT,`Region`	TEXT,`Admission`	TEXT,`Statut`	TEXT,`Points`	INTEGER);")

    for i in range(3,r):
        t2=[]
        for j in range(1, c+1):
            temp=sheet.cell(row=i, column=j).value
            t2.append(temp)
        t2=pointoecole(t2,Points)
        curseur.execute("INSERT INTO EcoleS (Acronyme,Nom,Commune,Region,Admission,Statut,Points,Groupe) VALUES  (?,?,?,?,?,?,?,?)",(t2[0],t2[1],t2[2],t2[3],t2[5],t2[6],t2[-1],t2[5]))

    connexion.commit()

def point():

    book2= openpyxl.load_workbook('sam.xlsx')
    concours=[]
    Fini=False

    while Fini==False:
        """On récupere les barres et les groupes associées dans l'excel"""

        debut=int(input("donne la ligne du début "))
        fin=int(input("donne la ligne de fin "))
        feuille=input("nom de la feuille ")
        colone=int(input("donne le numero de la colonne avec les barres "))
        sheet2=book2.get_sheet_by_name(feuille)

        for i in range(debut,fin+1):
            temp=sheet2.cell(row=i,column=1).value
            temp2=sheet2.cell(row=i,column=colone).value
            if temp2=="":
                temp2=9999
            concours.append([temp,temp2])
        Test=input("avez vous fini avec les barres des groupes alors dites oui sinon appuyer sur une touche et entree'")
        if Test=="oui":
            Fini=True
    return concours

Points=point()

def pointoecole(Ecole,Points):

    "On va itierer sur la base de données et d'abord rechercher par acroyme si l'école a sa propre barre sinon par nom "
    n=len(Ecole)
    for j in range(len(Points)):
        if Ecole[0]==Points[j][0]:
            Ecole.append(Points[j][1])

    if len(Ecole)!=n+1:

        for k in range(len(Points)):
            if Ecole[5]==Points[k][0]:
                Ecole.append(Points[k][1])

    if len(Ecole)!=n+1:
        Ecole.append(99999)
    return Ecole


def spe():

    curseur.execute("DROP TABLE Specialite")
    connexion.commit()
    curseur.execute("CREATE TABLE Specialite(Idspecialite INTEGER PRIMARY KEY AUTOINCREMENT,NomSpe TEXT)")
    for i in range(11,35):
        temp=sheet.cell(row=2,column=i).value
        curseur.execute("INSERT INTO Specialite (NomSpe) VALUES (?)",(temp,))


def alter():

    curseur.execute("DROP TABLE EcoleSpe")
    connexion.commit()
    curseur.execute(
        "CREATE TABLE `EcoleSpe` (`IdEcole`	INTEGER,`IdSpe` INTEGER,`Alternance` TEXT,FOREIGN KEY(`IdEcole`) REFERENCES `EcoleS`(`Id`),UNIQUE(`IdEcole`,`IdSpe`,`Alternance`),FOREIGN KEY(`IdSpe`) REFERENCES `Specialite`(`Idspecialite`));")

    for i in range(3, r):
        for j in range(11, 35):
            temp = sheet.cell(row=i, column=38).value
            temp2 = sheet.cell(row=i, column=j).value
            if temp2 == "OUI":
                temp = temp.capitalize()
                curseur.execute("INSERT INTO EcoleSpe (Alternance,IdSpe,IdEcole) VALUES (?,?,?)", (temp, j - 10, i - 2))
def creeDB():

    ecole()
    spe()
    alter()
    connexion.commit()

creeDB()