#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Oct 20 19:41:49 2018
@author: samir
"""

import sqlite3
import openpyxl



book = openpyxl.load_workbook("sam.xlsx")
sheet = book["S1"]
r = sheet.max_row

def connec(txt):
    global connexion, curseur

    connexion = sqlite3.connect(txt)
    curseur = connexion.cursor()


def renvoie_admission():

    return [
        resultat[0]
        for resultat in curseur.execute("SELECT DISTINCT Admission FROM EcoleS")
    ]


def moyenneel(nom):
    """Récupere les notes du canditat si son prenom/nom est dans l'excel sinon retourne des 0 partout """

    
    for i in range(4, r):
        temp = sheet.cell(row=i, column=1).value
        if temp == nom:
            note = [
                sheet.cell(row=i, column=4).value,
                sheet.cell(row=i, column=5).value,
                sheet.cell(row=i, column=6).value,
                sheet.cell(row=i, column=7).value,
                sheet.cell(row=i, column=8).value,
                sheet.cell(row=i, column=9).value,
                sheet.cell(row=i, column=10).value,
            ]
            return note
    return [0, 0, 0, 0, 0, 0]


def renvoie_specialites():
    return [
        resultat[0]
        for resultat in curseur.execute("SELECT DISTINCT NomSpe FROM Specialite")
    ]


def renvoie_regions():
    return [
        resultat[0]
        for resultat in curseur.execute("SELECT DISTINCT Region FROM EcoleS")
    ]


def prix_ecole(ecoles, filtre):
    prix = 0
    for i in list(set(ecoles)):

        for resultat in curseur.execute(
            "SELECT " + filtre + " FROM Coefficient WHERE Groupe=" + "'" + i + "'"
        ):
            prix += resultat[0]

    return prix


def renvoie_idspe(choix):
    
    return tuple(
        spe[0]
        for i in tuple(choix)
        for spe in curseur.execute(
            "SELECT idspecialite FROM Specialite WHERE NomSpe=" + "'" + i + "'"
        )
    )


def creationtuple(liste):

    if len(liste) == 1:
        return "('" + str(liste[0]) + "')"
   
    return tuple(liste)


def filtre(choix_utilisateur, notes):
    conds = []

    """Le none correspond au fait que l'utilisateur n'as rien choisi"""
    if choix_utilisateur["specialites"] != None:
        conds.append(["Idspe", "IN", choix_utilisateur["specialites"]])
    if choix_utilisateur["alternance"] != None:
        conds.append(["Alternance", "IN", choix_utilisateur["alternance"]])
    if choix_utilisateur["concours"] != None:
        conds.append(["Admission", "IN", choix_utilisateur["concours"]])
    if choix_utilisateur["regions"] != None:
        conds.append(["Region", "IN", choix_utilisateur["regions"]])

    if choix_utilisateur["annee"] == ("3/2",):
        bonif_str = "Bonification"
    else:
        bonif_str = "0"

    requete = (
        """
        SELECT DISTINCT id,Nom,Admission,Commune,Alternance,Acronyme,NomSpe
        FROM EcoleSpe
        JOIN EcoleS on EcoleSpe.IdEcole=EcoleS.id
        JOIN Specialite on EcoleSpe.IdSpe=Specialite.idspecialite
        JOIN Coefficient on Coefficient.Groupe=EcoleS.Groupe
        WHERE  """
        + str(notes["maths"])
        + "*Maths+"
        + str(notes["physique"])
        + "*Physique+"
        + str(notes["si"])
        + """*SI+"""
        + str(notes["informatique"])
        + "*Informatique+"
        + str(notes["anglais"])
        + "*Anglais+"
        + str(notes["francais"])
        + "*Francais+"
        + str(notes["modelisation"])
        + "*Modelisation+"
        + bonif_str
        + ">= Points   "
    )

    for var in conds:
        requete += " AND " + var[0] + "  " + var[1] + " " + str(creationtuple(var[2]))

 
    return [ecole for ecole in curseur.execute(requete)]
