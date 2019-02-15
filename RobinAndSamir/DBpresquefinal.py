#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 12 14:19:23 2019
@author: robinchaussemy
"""
import openpyxl
import sqlite3

connexion = sqlite3.connect("choixecole.db")
curseur = connexion.cursor()


""" Retourne la taille du fichier et les valeurs"""
    
book = openpyxl.load_workbook('Classeur1.xlsx')
sheet = book.get_sheet_by_name('Feuil1')
r = sheet.max_row
c = sheet.max_column

def ecole():
    curseur.execute("DROP TABLE EcoleS")
    connexion.commit()
    curseur.execute("CREATE TABLE EcoleS(Id INTEGER PRIMARY KEY AUTOINCREMENT,Acronyme TEXT,Nom	 TEXT,Groupe TEXT,Commune TEXT,CommuneAnnexe TEXT,Region TEXT,Admission TEXT,Statut 	TEXT,Points 	INTEGER,PrixB	INTEGER,PrixNB INTEGER)")

    for i in range(3,r):
        t2=[]
        for j in range(1, c+1):
            temp=sheet.cell(row=i, column=j).value
            t2.append(temp)
        curseur.execute("INSERT INTO EcoleS (Acronyme,Nom,Groupe,Commune,CommuneAnnexe,Region,Admission,Statut,Points,PrixB,PrixNB) VALUES  (?,?,?,?,?,?,?,?,?,?,?)",(t2[0],t2[1],"?",t2[2],t2[4],t2[3],t2[5],t2[6],0,t2[38],t2[39]))
    connexion.commit()
    
def spe():
    curseur.execute("DROP TABLE Specialite")
    connexion.commit()
    curseur.execute("CREATE TABLE Specialite(Id	 INTEGER PRIMARY KEY AUTOINCREMENT,NomSpe TEXT)")
    for i in range(11,35):
        temp=sheet.cell(row=2,column=i).value
        curseur.execute("INSERT INTO Specialite (NomSpe) VALUES (?)",(temp,))

def alter():
    curseur.execute("DROP TABLE EcoleSpe")
    connexion.commit()
    curseur.execute("CREATE TABLE `EcoleSpe` (`IdEcole`	INTEGER,`IdSpe` INTEGER,`Alternance` TEXT,FOREIGN KEY(`IdEcole`) REFERENCES `EcoleS`(`Id`),UNIQUE(`IdEcole`,`IdSpe`,`Alternance`),FOREIGN KEY(`IdSpe`) REFERENCES `Specialite`(`Id`));")
    
    for i in range(3,r):
        for j in range(11, 35):
            temp=sheet.cell(row=i,column=38).value
            temp2=sheet.cell(row=i, column=j).value
            if temp2=="OUI":
                temp=temp[0].upper()+temp[1::].lower()
                curseur.execute("INSERT INTO EcoleSpe (Alternance,IdSpe,IdEcole) VALUES (?,?,?)",(temp,j-10,i-2))
    
       
ecole()
spe()
alter()
connexion.commit()
