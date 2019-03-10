#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 12 14:19:23 2019
@author: robinchaussemy
"""
import openpyxl
import sqlite3

connexion = sqlite3.connect("choixecole.db")
curseur = connexion.cursor()


""" Retourne la taille du fichier et les valeurs"""
    
book = openpyxl.load_workbook('Classeur1.xlsx')
sheet = book.get_sheet_by_name('Feuil1')
book2= openpyxl.load_workbook('Barresadmissibilité.xlsx')
sheet2 =book2.get_sheet_by_name('Places CCINP')
book3 = openpyxl.load_workbook('Barresadmissibilité.xlsx')
sheet3 = book2.get_sheet_by_name('Places CS')
r = sheet.max_row
c = sheet.max_column

def ecole():
    curseur.execute("DROP TABLE EcoleS")
    connexion.commit()
    curseur.execute("CREATE TABLE EcoleS(Id	 INTEGER PRIMARY KEY AUTOINCREMENT,Acronyme 	TEXT,Nom	 TEXT,Groupe TEXT,Commune TEXT,CommuneAnnexe TEXT,Region TEXT,Admission TEXT,Statut 	TEXT,Points 	INTEGER,PrixB	INTEGER,PrixNB INTEGER)")

    for i in range(3,r):
        t2=[]
        for j in range(1, c+1):
            temp=sheet.cell(row=i, column=j).value
            t2.append(temp)
        curseur.execute("INSERT INTO EcoleS (Acronyme,Nom,Commune,CommuneAnnexe,Region,Admission,Statut,Points,PrixB,PrixNB,Groupe) VALUES  (?,?,?,?,?,?,?,?,?,?,?)",(t2[0],t2[1],t2[2],t2[4],t2[3],t2[5],t2[6],0,t2[38],t2[39],"?"))

    
   
    while 1 :
        
        nomdugroupe=input("Ecris le nom du groupe ")
        
        cdebut=int(input("Donne la colonne du début "))
        cfin=int(input("Donne la colonne du fin "))
        
        for i in range(cdebut,cfin+1):
            temp=sheet2.cell(row=i, column=4).value
            curseur.execute("UPDATE EcoleS SET Groupe = "+"'"+nomdugroupe+"'"+" WHERE Acronyme LIKE (?)",('%'+temp,))
        
        test=input("D'autres groupes a ajouter : ")
        
        if test.capitalize()!="Oui":
            break


    curseur.execute("UPDATE EcoleS SET Groupe = 'CESI' WHERE Admission LIKE 'Concours CESI'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'Cefipa' WHERE Admission LIKE 'Concours Cefipa'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'CCS' WHERE Admission LIKE 'Concours CS' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'Dossier et entretien' WHERE Admission LIKE 'Dossier et entretien' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'MinesPonts' WHERE Admission LIKE 'Mines-Ponts (CS)' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'MinesTélécom' WHERE Admission LIKE 'Mines-Télécom (CS)' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'BanqueépreuvesCCINP' WHERE Admission LIKE 'Banque CCINP' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'Epita/Ipsa' WHERE Admission LIKE 'Concours Epita/Ipsa' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'CCINP' WHERE Admission LIKE 'Concours CCINP' AND  Groupe = '?'")
    curseur.execute("UPDATE EcoleS SET Groupe = 'Banque CCS' WHERE Admission LIKE 'Banque CS' AND  Groupe = '?'")

    connexion.commit()

def spe():
    curseur.execute("DROP TABLE Specialite")
    connexion.commit()
    curseur.execute("CREATE TABLE Specialite(Idspecialite INTEGER PRIMARY KEY AUTOINCREMENT,NomSpe TEXT)")
    for i in range(11,35):
        temp=sheet.cell(row=2,column=i).value
        curseur.execute("INSERT INTO Specialite (NomSpe) VALUES (?)",(temp,))


def alter():
    curseur.execute("DROP TABLE EcoleSpe")
    connexion.commit()
    curseur.execute(
        "CREATE TABLE `EcoleSpe` (`IdEcole`	INTEGER,`IdSpe` INTEGER,`Alternance` TEXT,FOREIGN KEY(`IdEcole`) REFERENCES `EcoleS`(`Id`),UNIQUE(`IdEcole`,`IdSpe`,`Alternance`),FOREIGN KEY(`IdSpe`) REFERENCES `Specialite`(`Id`));")

    for i in range(3, r):
        for j in range(11, 35):
            temp = sheet.cell(row=i, column=38).value
            temp2 = sheet.cell(row=i, column=j).value
            if temp2 == "OUI":
                temp = temp.capitalize()
                curseur.execute("INSERT INTO EcoleSpe (Alternance,IdSpe,IdEcole) VALUES (?,?,?)", (temp, j - 10, i - 2))
def creeDB():

    ecole()
    spe()
    alter()
    connexion.commit()

creeDB()