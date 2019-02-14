#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 12 14:19:23 2019

@author: robinchaussemy
"""
import openpyxl
import sqlite3

connexion = sqlite3.connect("choixecole1.db")
curseur = connexion.cursor()


""" Retourne la taille du fichier et les valeurs"""
    
book = openpyxl.load_workbook('Classeur1.xlsx')
sheet = book.get_sheet_by_name('Feuil1')
r = sheet.max_row
c = sheet.max_column

def ecole():
    for i in range(3,r):
        t2=[]
        for j in range(1, c+1):
            temp=sheet.cell(row=i, column=j).value
            t2.append(temp)
        curseur.execute("INSERT INTO EcoleS (Acronyme,Nom,Groupe,Commune,CommuneAnnexe,Region,Admission,Statut,Points,PrixB,PrixNB) VALUES  (?,?,?,?,?,?,?,?,?,?,?)",(t2[0],t2[1],"?",t2[2],t2[4],t2[3],t2[5],t2[6],0,t2[38],t2[39]))
    connexion.commit()
    
def spe():
    for i in range(11,35):
        temp=sheet.cell(row=2,column=i).value
        curseur.execute("INSERT INTO Specialite (Nomspe) VALUES (?)",(temp,))

    
       
spe()
connexion.commit()