#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Oct 20 19:41:49 2018
@author: samir
"""

import sqlite3
import openpyxl

book = openpyxl.load_workbook('Barresadmissibilité.xlsx')
sheet = book.get_sheet_by_name('S1')

connexion = sqlite3.connect("choixecole.db")
curseur = connexion.cursor()


def renvoie_admission():
    return [resultat[0] for resultat in curseur.execute("SELECT DISTINCT Admission FROM EcoleS")]
def moyenneel(nom):
    r = sheet.max_row
    for i in range(4,r):
        temp = sheet.cell(row=i, column=1).value
        if temp == nom:
            note=[sheet.cell(row=i, column=4).value,sheet.cell(row=i, column=5).value,sheet.cell(row=i, column=6).value,sheet.cell(row=i, column=7).value,sheet.cell(row=i, column=8).value,sheet.cell(row=i, column=9).value,sheet.cell(row=i, column=10).value]
    return(note)
def renvoie_specialites():
    return [resultat[0] for resultat in curseur.execute("SELECT DISTINCT NomSpe FROM Specialite")]


def renvoie_regions():
    return [resultat[0] for resultat in curseur.execute("SELECT DISTINCT Region FROM EcoleS")]


def renvoie_idspe(choix):
    idspe = tuple()
    for i in tuple(choix):
        idspe += tuple(
            spe[0] for spe in curseur.execute("SELECT idspecialite FROM Specialite WHERE NomSpe=" + "'" + i + "'"))
    return idspe


def creationtuple(liste):
    variablein = ""

    for i in range(len(liste) - 1):
        variablein += "'" + str(liste[i]) + "',"

    return variablein + "'" + str(liste[-1]) + "'"


def filtre(choix_utilisateur, notes):
    conds = []

    if choix_utilisateur["specialites"] != None:
        conds.append(["Idspe", "IN", choix_utilisateur["specialites"]])
    if choix_utilisateur["alternance"] != None:
        conds.append(["Alternance", "IN", choix_utilisateur["alternance"]])
    if choix_utilisateur["concours"] != None:
        conds.append(["Admission", "IN", choix_utilisateur["concours"]])
    if choix_utilisateur["regions"] != None:
        conds.append(["Region", "IN", choix_utilisateur["regions"]])

    if choix_utilisateur["annee"] == ("3/2",):
        bonif_str = "Bonification"
    else:
        bonif_str = "0"

    requete = ("""
        SELECT DISTINCT id,Nom,Admission,Commune,Alternance,Acronyme,NomSpe
        FROM EcoleSpe
        JOIN EcoleS on EcoleSpe.IdEcole=EcoleS.id
        JOIN Specialite on EcoleSpe.IdSpe=Specialite.idspecialite
        JOIN Coefficient on Coefficient.Groupe=EcoleS.Groupe
        WHERE """
               + str(notes["maths"]) + "*Maths+"
               + str(notes["physique"]) + "*Physique+"
               + str(notes["si"]) + """*SI+"""
               + str(notes["informatique"]) + "*Informatique+"
               + str(notes["anglais"]) + "*Anglais+"
               + str(notes["francais"]) + "*Francais+"
               + str(notes["modelisation"]) + "*Modelisation+"
               + bonif_str
               + ">= Points  "

               )

    for var in conds:
        variablein = creationtuple(var[2])
        requete += " AND " + var[0] + " " + var[1] + " (" + variablein + ")"
    print(notes)

    return [ecole for ecole in curseur.execute(requete)]